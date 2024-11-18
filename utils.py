import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import os
import cv2
import datetime as dt
import re
import math
import openpyxl
from torch.utils.data import DataLoader
from scipy import interpolate

from dataset import NNDataset, collate_fn_custom

def get_match_idx_coord(idx_neural, idx_coord_neural):
    '''
    Find coord index that matches to neural index for the first time
    idx_neural: int index
    '''        
    idx_list = np.where(idx_coord_neural == idx_neural)[0]
    if len(idx_list) == 0:
        print('No matching idx_coord_neural, in get_match_idx_coord()')
        exit()
    else:
        return idx_list[0]

def get_match_idx_only(idx_coord_neural):
    '''
    return neural indices that only match to coord and coord indices that only have matched neural idx
    '''
    idx_coord_neural_nonan = [x for x in idx_coord_neural if math.isnan(x) is False]
    idx_coord_match = []
    for i in set(idx_coord_neural_nonan):
        itemindex = get_match_idx_coord(i, idx_coord_neural)
        idx_coord_match.append(itemindex)
    idx_neural_match = [int(idx_coord_neural[i]) for i in idx_coord_match]

    return idx_neural_match, idx_coord_match

def load_coord_csv(path_csv, animal_name):
    
    if animal_name == 'Animal1-G8_53950_1L_Redo' or animal_name == 'Animal3-G8_53950_2R_Redo':
        behavior_coord = pd.DataFrame(pd.read_csv(path_csv, header=[2]))
        fx = np.array([behavior_coord.loc[:, 'x'].to_numpy()])
        fy = np.array([behavior_coord.loc[:, 'y'].to_numpy()])
        hx = np.array([behavior_coord.loc[:, 'x.1'].to_numpy()])
        hy = np.array([behavior_coord.loc[:, 'y.1'].to_numpy()])
        fl = np.array([behavior_coord.loc[:, 'likelihood'].to_numpy()])
        hl = np.array([behavior_coord.loc[:, 'likelihood.1'].to_numpy()])

    elif animal_name == "Animal2-G16_55875_1L_2nd":
        behavior_coord = pd.DataFrame(pd.read_csv(path_csv, header=[1,2]))
        fx = np.array([(behavior_coord.loc[:, 'FR_paw']).loc[:,'x'].to_numpy()])
        fy = np.array([(behavior_coord.loc[:, 'FR_paw']).loc[:,'y'].to_numpy()])
        hx = np.array([(behavior_coord.loc[:, 'HR_paw']).loc[:,'x'].to_numpy()])
        hy = np.array([(behavior_coord.loc[:, 'HR_paw']).loc[:,'y'].to_numpy()])
        fl = np.array([(behavior_coord.loc[:, 'FR_paw']).loc[:,'likelihood'].to_numpy()])
        hl = np.array([(behavior_coord.loc[:, 'HR_paw']).loc[:,'likelihood'].to_numpy()])

    elif animal_name == "Animal4-G10_55902_1R" or animal_name == "Animal5-G10_55903_1R" or animal_name == "Animal6-G10_55904_1L" or \
        animal_name == "Animal7-G12_55946_1R" or animal_name == "Animal8-G12_55947_1R" or animal_name == "Animal9-G12_55954_1L":
        behavior_coord = pd.DataFrame(pd.read_csv(path_csv, header=[1,2]))
        fx = np.array([(behavior_coord.loc[:, 'F_paw']).loc[:,'x'].to_numpy()])
        fy = np.array([(behavior_coord.loc[:, 'F_paw']).loc[:,'y'].to_numpy()])
        hx = np.array([(behavior_coord.loc[:, 'H_paw']).loc[:,'x'].to_numpy()])
        hy = np.array([(behavior_coord.loc[:, 'H_paw']).loc[:,'y'].to_numpy()])
        fl = np.array([(behavior_coord.loc[:, 'F_paw']).loc[:,'likelihood'].to_numpy()])
        hl = np.array([(behavior_coord.loc[:, 'H_paw']).loc[:,'likelihood'].to_numpy()])

    else:
        print("No matched animal name")
        print(animal_name)
        exit()

    behavior_coord = np.concatenate((fx,fy,hx,hy))
    likeli = np.concatenate((fl,hl))

    return behavior_coord, likeli

def load_video_cv2(path_video, frame_range=None, frame_idx=None):
    '''
    https://stackoverflow.com/questions/42163058/how-to-turn-a-video-into-numpy-array

    frame_range = [start_idx, end_idx]. include end_idx
    frame_idx = int

    '''
    # print('Loading video')

    cap = cv2.VideoCapture(path_video)

    fps = cap.get(cv2.CAP_PROP_FPS)
    # print('frames per second =',fps)

    total_num_frame = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    # print('total number of frames: ', total_num_frame)

    if frame_range is not None:
        cap.set(cv2.CAP_PROP_POS_FRAMES, frame_range[0])
        frameCount = frame_range[1] - frame_range[0] +1 
    else:
        frameCount = total_num_frame
    
    if frame_idx is not None:
        cap.set(cv2.CAP_PROP_POS_FRAMES, frame_idx)
        frameCount = 1

    frameWidth = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    frameHeight = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    buf = np.empty((frameCount, frameHeight, frameWidth, 3), np.dtype('uint8'))
    fc = 0
    ret = True
    # print("Start reading frame")
    while (fc < frameCount  and ret):
        # if fc % 5000 == 0 or fc == frameCount-1:
            # print(fc, '/', frameCount)
        ret, buf[fc] = cap.read()
        fc += 1
        # if fc % 10000 == 0 or fc == frameCount-1:
    cap.release()

    return buf

def get_video_len(path_video):
    cap = cv2.VideoCapture(path_video)
    total_num_frame = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    # print('total number of frames: ', total_num_frame)

    return total_num_frame

def get_name_with_time(filename):
    timestamp = str(dt.datetime.now())[:19]
    timestamp = re.sub(r'[\:-]','', timestamp) # replace unwanted chars 
    timestamp = re.sub(r'[\s]','_', timestamp) # with regex and re.sub
    out_filename = ('{}_{}'.format(timestamp,filename))

    return out_filename

def convert_norm_coord_func(arr_norm, arr_ori):
    if arr_norm.shape[0] != arr_ori.shape[0]:
        print('Shape[0] is different. ', arr_norm.shape, arr_ori.shape)
        exit()

    convert_arr = np.zeros(arr_norm.shape)
    for i in range(arr_norm.shape[0]):
        row = arr_ori[i, :]
        max_row = max(row)
        min_row = min(row)
        convert = arr_norm[i,:]*(max_row-min_row) + min_row
        convert = np.round(convert, 6)

        convert_arr[i,:] = convert
    
    # checking with gt
    # if np.array_equal(convert_arr, arr_ori) is False:
    #     print('converted arr is different.')
    #     print('arr_ori: ', arr_ori[0,:10])
    #     print('convert_arr: ', convert_arr[0,:10])
    #     exit()
    # if np.allclose(convert_arr, arr_ori, rtol=1e-05) is False:
    #     print('ori and convert is different with rtol=1e-05.')
    #     print('arr_ori: ', arr_ori[0,:10])
    #     print('convert_arr: ', convert_arr[0,:10])
    #     exit()
    # else:
    #     print('ori and convert is close with rtol=1e-05.')
    
    return convert_arr

# def get_dir_data_from_dir_result(dir_result):
#     path_txt = os.path.join(dir_result, 'log.txt')
#     with open(path_txt, 'r') as f:
#         lines = f.readlines()
#         dir_data_line = lines[2]
#         if dir_data_line[:10] != "dir_data :":
#             print('get_dir_data_from_dir_result is wrong.')
#             print(dir_data_line[:10])
#             exit()
        
#         return dir_data_line[10:-1]

class ExcelSaver():
    def __init__(self, path_excel, list_col=None, sheet_type='single', sheet_name=None):
        '''
        sheet_type: 'multi','single'
        list_col: [name1, name2, ...] if 'single', [[],[],...] if 'multi'
        sheet_name: name if 'single', [name1, name2, ...] if 'multi'
        '''
        
        if sheet_type == 'multi' and sheet_name is None:
            print('Give sheet_name in ExcelSaver')
            exit()
        
        self.path_excel = path_excel
        wb = openpyxl.Workbook()
        if sheet_type == 'single':
            if list_col is not None:
                ws = wb.active
                ws.append(list_col)
        elif sheet_type == 'multi':
            for i in range(len(sheet_name)):
                if i == 0:
                    ws = wb['Sheet']
                    ws.title = sheet_name[i]
                else:
                    ws = wb.create_sheet(sheet_name[i])
                if list_col is not None: ws.append(list_col[i])
        wb.save(path_excel)
            
    def save_row(self, data_list, ws_name=None):
        wb = openpyxl.load_workbook(self.path_excel)
        ws = None
        if ws_name is None: ws = wb.active
        else: ws = wb[ws_name]
        ws.append(data_list)
        wb.save(self.path_excel)

def read_excel(path_excel):
    '''
    return col_names, dict_excel
    '''
    df = pd.read_excel(path_excel)
    col_names = list(df.columns)
    dict_excel = {}
    for i in col_names:
        dict_excel[i] = df.loc[:,i].tolist()
    print('col_names: ', col_names)
    return col_names, dict_excel

def get_coord_idx_stack(idx_neuron, DIR_DATA, name_neural, name_coord, seq_len, output_idx, dataset, SIZE_BATCH_VAL, gt_load=None):
    '''
    get coord_idx_stack, which is idx_coord_neural of test dataset
    '''
    # Load gt_stack and coord_idx using dataset
    test_dataset = NNDataset(DIR_DATA, name_neural, name_coord, seq_len, output_idx, dataset, idx_neuron)
    dataloader_test = DataLoader(test_dataset, batch_size=SIZE_BATCH_VAL, pin_memory=True, shuffle=False, collate_fn=collate_fn_custom)

    gt_stack = None
    coord_idx_stack = []
    for idx_batch, data in enumerate(dataloader_test):
        # Load data
        inputs, labels, _, coord_idx = data
        # print('inputs: ', inputs.shape)
        #print('labels: ', labels.shape)
        # print('coord_idx: ', len(coord_idx))

        labels_reshape = None
        if labels.dim() == 3: labels_reshape = np.reshape(labels.data.cpu().numpy(), (labels.shape[0]*labels.shape[1], labels.shape[2]))
        elif labels.dim() == 2: labels_reshape = labels.data.cpu().numpy()
        #print('labels_reshape: ', labels_reshape.shape)

        if gt_stack is None: gt_stack = labels_reshape
        else: gt_stack = np.concatenate((gt_stack, labels_reshape))

        for i in coord_idx:
            coord_idx_stack.extend(i)

    # reverse mask, transpose
    idx = np.any(gt_stack, axis=1)
    gt_stack = gt_stack[idx, :]
    gt_stack = np.transpose(gt_stack)
    #print('gt_stack: ', gt_stack.shape)
    #print('coord_idx_stack: ', len(coord_idx_stack), coord_idx_stack[:10], coord_idx_stack[-10:])

    # Reverse norm convert coord for gt_stack if behav_coord_norm
    if 'norm' in name_coord:
        ori = np.load(os.path.join(DIR_DATA, 'behav_coord_ori.npy'), allow_pickle=True)
        convert_arr = np.zeros(gt_stack.shape)
        for i in range(gt_stack.shape[0]):
            row = ori[i, :]
            max_row = max(row)
            min_row = min(row)
            convert = gt_stack[i,:]*(max_row-min_row) + min_row
            convert = np.round(convert, 5)
            convert_arr[i,:] = convert
        gt_stack = convert_arr

    # Compare gt from dataset and model result directory
    if gt_load is not None:
        flag_equal = np.array_equal(gt_stack, gt_load)
        #print('Equal gt_stack vs gt: ', flag_equal)


        # print('gt_stack: ', convert_arr[0,:10])
        # print('gt_output: ', gt[0,:10])
        # plt.plot(convert_arr[0,:], label='gt_stack')
        # plt.plot(gt[0,:], label='gt_output', alpha=0.5)
        # plt.legend()
        # plt.show()

        if flag_equal is False:
            exit()

    return coord_idx_stack, gt_stack

def interp_xy_1d(dim1, dim2, len_win, num_interp):
    '''
    https://stackoverflow.com/questions/52014197/how-to-interpolate-a-2d-curve-in-python
    '''
    # len_win = 5
    # interp_multiple = 3
    
    dim1_interp, dim2_interp = [], []
    for idx_win in range(dim1.shape[0]//len_win):
        if idx_win == dim1.shape[0]//len_win-1:  
            dim1_win = dim1[idx_win*len_win:]
            dim2_win = dim2[idx_win*len_win:]
        else: 
            dim1_win = dim1[idx_win*len_win:(idx_win+1)*len_win]
            dim2_win = dim2[idx_win*len_win:(idx_win+1)*len_win]

        points = np.array([dim1_win, dim2_win]).T #(nbre_points x nbre_dim)
        #print('points: ', points.shape)

        distance = np.cumsum( np.sqrt(np.sum( np.diff(points, axis=0)**2, axis=1 )) )
        distance = np.insert(distance, 0, 0)/distance[-1]

        method = 'quadratic'
        alpha = np.linspace(0, 1, dim1_win.shape[0]*num_interp)
        interpolator =  interpolate.interp1d(distance, points, kind=method, axis=0)
        interpolated_points = interpolator(alpha)
        #print('interpolated_points: ', interpolated_points.shape) #(1870, 2)

        dim1_interp.extend(interpolated_points[:,0])
        dim2_interp.extend(interpolated_points[:,1])
    
    return dim1_interp, dim2_interp

def get_idx_coord_neural_of_test(dir_data):
    '''
    get idx_coord_neural of test data
    '''
    test_idx_cv = np.load(os.path.join(dir_data, 'test_idx_cv.npy'))
    test_idx = test_idx_cv[-1]
    #print('test_idx: ', test_idx)
    
    idx_coord_neural = np.load(os.path.join(dir_data, 'idx_coord_neural.npy'))

    idx_coord_start = np.where(idx_coord_neural == test_idx[0])[0][0]
    #print(idx_coord_start, idx_coord_neural[idx_coord_start])

    idx_coord_end = np.where(idx_coord_neural == test_idx[1])[0][-1]
    #print(idx_coord_end, idx_coord_neural[idx_coord_end])

    idx_coord_neural_test = idx_coord_neural[idx_coord_start:idx_coord_end+1]
    #print('idx_coord_neural_test: ', idx_coord_neural_test[:5], idx_coord_neural_test[-5:])

    return idx_coord_neural_test

def get_run_stand_wins_coord(coord, flag_figure, limb_to_crop_str=None, dir_fig_save=None, title=None):
    '''
    Get run and stand windows of coord
    coord: (limb, seq_len)
    limb_to_crop_str: 'RFX', 'RFY', 'RHX', 'RHY', 'LFX', 'LFY', 'LHX', 'LHY'
    '''

    # if limb_to_crop is None:
    #     if coord.shape[0] == 8:
    #         limb_to_crop = 4
    #     elif coord.shape[0] == 4:
    #         limb_to_crop = 2
    #     else:
    #         limb_to_crop = 0
    limb_to_crop = None
    if coord.shape[0] == 8:
        match limb_to_crop_str:
            case 'RFX': limb_to_crop = 0
            case 'RFY': limb_to_crop = 1
            case 'RHX': limb_to_crop = 2
            case 'RHY': limb_to_crop = 3
            case 'LFX': limb_to_crop = 4
            case 'LFY': limb_to_crop = 5
            case 'LHX': limb_to_crop = 6
            case 'LHY': limb_to_crop = 7
    elif coord.shape[0] == 4:
        match limb_to_crop_str:
            case 'RFX': limb_to_crop = 0
            case 'RFY': limb_to_crop = 0
            case 'RHX': limb_to_crop = 1
            case 'RHY': limb_to_crop = 1
            case 'LFX': limb_to_crop = 2
            case 'LFY': limb_to_crop = 2
            case 'LHX': limb_to_crop = 3
            case 'LHY': limb_to_crop = 3
    else:
        print('is coord right? in get_run_stand_wins_coord')
        exit()

    thrd_pixel = 10 #thrd for displacement 
    thrd_frame = 30 #minimum length of run period
    margin_continue = 30 #maximum interval length to judge as continued period  

    # Get idx_higher
    gt_limb_to_crop = coord[limb_to_crop, :]
    gt_limb_to_crop_dif = [abs(gt_limb_to_crop[i]-gt_limb_to_crop[i-1]) for i in range(1,gt_limb_to_crop.shape[0])]
    idx_higher = [i for i in range(1,gt_limb_to_crop.shape[0]) if gt_limb_to_crop_dif[i-1]>thrd_pixel]
    
    # Get periods that meet the conditions
    run_windows = []
    prev = None
    tmp = []
    if idx_higher[0] < thrd_frame:
        tmp.extend([0])
        prev = 0
    for i in idx_higher:
        if prev is None or i-prev < margin_continue:
            tmp.append(i)
        else:
            if len(tmp)>thrd_frame: #if len(tmp)!=0
                run_windows.append([tmp[0], tmp[-1]])
                tmp = []
            else:
                tmp = []
                #tmp.append(i)
        prev = i
    if len(tmp) != 0: run_windows.append([tmp[0], tmp[-1]])
    print('run_windows: ', run_windows)
    run_windows = np.array(run_windows)
    print(run_windows.shape)

    ### Get stand periods
    stand_windows = []
    if run_windows[0][0] != 0:
        stand_windows.append([0, run_windows[0][0]-1])
    for i in range(len(run_windows)-1):
        run_win_left = run_windows[i]
        run_win_right = run_windows[i+1]
        stand_windows.append([run_win_left[-1]+1, run_win_right[0]-1])
    if run_windows[-1][-1] != coord.shape[1]-1:
        stand_windows.append([run_windows[-1][-1]+1, coord.shape[1]-1])
    print('stand_windows: ', stand_windows)
    stand_windows = np.array(stand_windows)
    print(stand_windows.shape)

    # Check 
    num_count = 0
    for i in run_windows:
        num_count += i[1]-i[0]+1
    for i in stand_windows:
        num_count += i[1]-i[0]+1
    if num_count != coord.shape[1]:
        print('num_count != coord.shape[1]')
        exit()
    # else:
    #     print('num_count = coord.shape[1]')
    
    if flag_figure:
        heatmap = np.zeros((coord.shape[0], coord.shape[1], 1))
        for i, win in enumerate(run_windows):
            for j in range(coord.shape[0]):
                heatmap[j, win[0]:win[1]+1,0] = 1
        
        for idx_limb in range(coord.shape[0]):
            coord_limb = coord[idx_limb, :]
            y_min = np.min(coord_limb)
            y_max = np.max(coord_limb)

            fig = plt.figure()
            plt.plot(coord_limb, color='black')
            plt.imshow(np.transpose(heatmap[idx_limb,:,:]), extent=[0, coord.shape[1], y_min, y_max], aspect='auto', cmap='Reds', vmin=0, vmax=1, interpolation='none')
            plt.colorbar()
            fig.set_size_inches(18,6)
            if dir_fig_save is None:
                plt.show()
            else: 
                if title is None:
                    plt.savefig(os.path.join(dir_fig_save, 'run_'+str(idx_limb)+'.png'))
                    plt.savefig(os.path.join(dir_fig_save, 'run_'+str(idx_limb)+'.pdf'))
                else:
                    plt.savefig(os.path.join(dir_fig_save, 'run_'+title+'_'+str(idx_limb)+'.png'))
                    plt.savefig(os.path.join(dir_fig_save, 'run_'+title+'_'+str(idx_limb)+'.pdf'))
            plt.close()

    return run_windows, stand_windows


def load_idx_coord_stack(dir_result, best_epoch):
    return np.load(os.path.join(dir_result, 'output', 'val', 'epoch_'+str(best_epoch)+'_idx_coord_stack.npy'), allow_pickle=True) 

def convert_idx_coord_real_to_test(dir_result, best_epoch, coord_idx_real_list):
    '''
    convert idx_coord_real to test indices
    '''
    # Get the idx_coord_test_start
    idx_coord_stack = load_idx_coord_stack(dir_result, best_epoch)
    idx_coord_test_start = idx_coord_stack[0]
    
    return [x-idx_coord_test_start for x in coord_idx_real_list]

def find_best_epoch_normconvert(dir_result):
    best_epoch = None
    for i in os.listdir(dir_result):
        if 'gt_norm_converted_epoch' in i:
            best_epoch = int(i[i.index('epoch')+5:i.index('.')])
    
    return best_epoch